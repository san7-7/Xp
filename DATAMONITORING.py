from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import openpyxl
from pushbullet import Pushbullet

# Set the path to the chromedriver executable
webdriver_path = 'C:/Users/SANRZza/Desktop/SANRZza/SANDocs/ProjectX/chromedriver-win64/chromedriver-win64/chromedriver.exe'

# Pushbullet API key
pushbullet_api_key = 'o.MV28DCeyoV7wjXx4lRyaeb2S2ucdD6Pv'
pb = Pushbullet(pushbullet_api_key)

# Create Chrome options
chrome_options = webdriver.ChromeOptions()

# Create the ChromeDriver service
chrome_service = Service(executable_path=webdriver_path)

# Create a new instance of the Chrome WebDriver with options
driver = webdriver.Chrome(service=chrome_service, options=chrome_options)

# Open the webpage
url = 'https://500.casino/wheel'
driver.get(url)

# CSS selector for the div elements
div_selector = 'div.round-history-list-item'

# Set the total duration to run the loop (60 minutes)
total_duration = 60 * 120  # 60 minutes in seconds

# Load the existing Excel file
existing_file_path = 'round_data.xlsx'
workbook = openpyxl.load_workbook(existing_file_path)
sheet = workbook.active

# Find the next empty row in the Excel sheet
next_empty_row = sheet.max_row + 1

# Set to store processed round numbers
processed_round_numbers = set()

# Counter for consecutive rounds of the same color
consecutive_color_count = 0

# Color to track
target_color = 'black'  # Replace with the color you want to track



# Variables to track the gap between two black
last_black_round = 0
rounds_since_last_black = 0

# Variables to track the gap between two blues
last_blue_round = 0
rounds_since_last_blue = 0
# variables to track the gap between two reds
last_red_round = 0
rounds_since_last_red = 0
# variables to track the gap between two gold
last_gold_round = 0
rounds_since_last_gold = 0

start_time = time.time()

def send_notification(message):
    pb.push_note("Color Notification", message)

try:
    while time.time() - start_time < total_duration:
        # Refresh the list of div elements with the class 'round-history-list-item'
        div_elements = driver.find_elements(By.CSS_SELECTOR, div_selector)

        # Iterate through each div element and extract round number and color
        for div_element in div_elements:
            try:
                # Extract round number and color
                round_number = div_element.get_attribute('to').split('-')[-1]
                color = div_element.get_attribute('class').split()[-1]

                # Check if this round number has been processed before
                if round_number not in processed_round_numbers:
                    print(f"Round Number: {round_number}\nColor: {color}\n")

                    # Add data to the Excel sheet
                    sheet.cell(row=next_empty_row, column=7, value=round_number)
                    sheet.cell(row=next_empty_row, column=8, value=color)
                    processed_round_numbers.add(round_number)

                    # Check if the current color is the target color
                    if color == target_color:
                        consecutive_color_count += 1
                        print(f"Consecutive {target_color} Count: {consecutive_color_count}")

                        # Check if the consecutive count is 7 or more
                        if consecutive_color_count == 7:
                            message = f"Consecutive {target_color} Count: {consecutive_color_count}"
                            send_notification(message)

                    else:
                        # Reset the consecutive count if the color changes
                        consecutive_color_count = 0

                    # Check if the current color is blue
                    if color == 'blue':
                        last_blue_round = int(round_number)
                        rounds_since_last_blue = 0

                    else:
                        # Increment rounds_since_last_blue if the color is not blue
                        rounds_since_last_blue += 1

                        # Check if rounds_since_last_blue exceeds 15
                        if rounds_since_last_blue >= 50:
                            message = f"More than 50,  {rounds_since_last_black} rounds since the last blue"
                            send_notification(message)
                        elif rounds_since_last_blue >= 40:
                            message = f"More than 40,  {rounds_since_last_black} rounds since the last blue"
                            send_notification(message)
                        elif rounds_since_last_blue >= 30:
                            message = f"More than 30,  {rounds_since_last_black} rounds since the last Blue"
                            send_notification(message)
                        elif rounds_since_last_blue >= 20:
                            message = f"More than 20, {rounds_since_last_black} rounds since the last BLue"
                            send_notification(message)
                        elif rounds_since_last_blue >= 15:
                            message = f"More than 15, {rounds_since_last_blue} rounds since the last BlUE"
                            send_notification(message)

                     # Check if the current color is black
                    if color == 'black':
                        last_black_round = int(round_number)
                        rounds_since_last_black = 0

                    else:
                        # Increment rounds_since_last_blue if the color is not blue
                        rounds_since_last_red += 1

                        # Check if rounds_since_last_blue exceeds 30
                        if rounds_since_last_black >= 17:
                            message = f"More than  17, {rounds_since_last_black} rounds since the last BLACK"
                            send_notification(message)
                        elif rounds_since_last_black >= 14:
                            message = f"More than 14,  {rounds_since_last_black} rounds since the last BLAck"
                            send_notification(message)
                        elif rounds_since_last_black >= 10:
                            message = f"More than eq 10, {rounds_since_last_black} rounds since the last BLack"
                            send_notification(message)
                        elif rounds_since_last_black == 8:
                            message = f" 8,  rounds since the last Black"
                            send_notification(message)
                        elif rounds_since_last_black == 5:
                            message = f" 5  rounds since the last black"
                            send_notification(message)

                    # Check if the current color is red
                    if color == 'red':
                        last_red_round = int(round_number)
                        rounds_since_last_red = 0

                    else:
                        # Increment rounds_since_last_blue if the color is not blue
                        rounds_since_last_red += 1

                        # Check if rounds_since_last_blue exceeds 30
                        if rounds_since_last_red >=30:
                            message = f"More than 30, {rounds_since_last_red} rounds since the last RED"
                            send_notification(message)
                        elif rounds_since_last_red >=25:
                            message = f"More than 25, {rounds_since_last_red} rounds since the last REd"
                            send_notification(message)
                        elif rounds_since_last_red >=20:
                            message = f"More than 20, {rounds_since_last_red} rounds since the last Red"
                            send_notification(message)
                        elif rounds_since_last_red >=15:
                            message = f"More than 15, {rounds_since_last_red}  rounds since the last red"
                            send_notification(message)
                        elif rounds_since_last_red >=12:
                            message = f"More than 12, {rounds_since_last_red} rounds since the last red"
                            send_notification(message)
                    

                    # Check if the current color is gold
                    if color == 'yellow':
                        last_gold_round = int(round_number)
                        rounds_since_last_gold = 0

                    else:
                        # Increment rounds_since_last_gold if the color is not gold
                        rounds_since_last_gold += 1

                        # Check if rounds_since_last_blue exceeds 30
                        if rounds_since_last_gold >= 400:
                            message = f"More than 400, {rounds_since_last_gold} rounds since the last GDamold"
                            send_notification(message)
                        elif rounds_since_last_gold >= 350:
                            message = f"More than 350, {rounds_since_last_gold} rounds since the last GOLD"
                            send_notification(message)
                        elif rounds_since_last_gold >= 300:
                            message = f"More than 300, {rounds_since_last_gold} rounds since the last GOLD"
                            send_notification(message)
                        elif rounds_since_last_gold >= 250:
                            message = f"More than 250, {rounds_since_last_gold} rounds since the last GOld"
                            send_notification(message)
                        elif rounds_since_last_gold >= 200:
                            message = f"More than 200, {rounds_since_last_gold} rounds since the last Gold"
                            send_notification(message)
                        elif rounds_since_last_gold >= 150:
                            message = f"More than 350, {rounds_since_last_gold} rounds since the last gold"
                            send_notification(message)
                        elif rounds_since_last_gold >= 100:
                            message = f"More than 100, {rounds_since_last_gold} rounds since the last gold"
                            send_notification(message)



                    # Increment the next empty row
                    next_empty_row += 1

            except Exception as e:
                print(f"Error extracting data from element: {e}")

        # Wait for 35 seconds before checking for new elements
        time.sleep(33)

except KeyboardInterrupt:
    # Handle KeyboardInterrupt (Ctrl+C) to exit the loop gracefully
    pass

finally:
    # Save the modified Excel file
    workbook.save(existing_file_path)

    # Close the browser window
    driver.quit()
